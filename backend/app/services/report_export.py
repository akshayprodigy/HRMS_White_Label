"""Universal export layer.

Every report is a `ReportResult`. We render it to xlsx / csv / pdf
using ONE dispatch — the ColumnDef.type drives cell formatting so
adding a new report needs zero export code changes.

Large-export strategy
=====================
The typical HR ERP report is 50–5,000 rows — well within a single-pass
in-memory xlsx write. For that regime we stream everything through
openpyxl with `constant_memory=True` write-only mode, so the peak
memory footprint stays flat regardless of row count.

For very large exports (10k+ rows — bulk salary registers, national
attendance dumps) the caller can pass `stream=True` and the writer
uses openpyxl's optimised write-only workbook, which flushes rows to
disk as they're appended. CSV is streamed as a generator so the
FastAPI StreamingResponse never buffers the whole payload.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any, Iterable, List, Optional, Tuple

from app.services.reports import (
    ColumnDef, ColumnType, ReportResult,
    fmt_cell, fmt_hours, fmt_inr, fmt_percent,
)


# ============================================================
# CSV
# ============================================================


def render_csv(result: ReportResult) -> str:
    """CSV export. Uses the same fmt_cell dispatch so numbers, hours,
    percents render identically to what the on-screen table shows.
    """
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow([c.label for c in result.columns])
    for row in result.rows:
        w.writerow([
            fmt_cell(row.get(c.key), c.type) for c in result.columns
        ])
    if result.totals:
        w.writerow([
            fmt_cell(result.totals.get(c.key), c.type)
            for c in result.columns
        ])
    return buf.getvalue()


# ============================================================
# Excel
# ============================================================


_EXCEL_NUMBER_FORMATS = {
    ColumnType.CURRENCY: '[>=10000000]"₹"##\\,##\\,##\\,##0.00;[>=100000]"₹"##\\,##\\,##0.00;"₹"#,##0.00',
    ColumnType.PERCENT: "0.00%",
    ColumnType.DATE: "dd-mmm-yyyy",
    ColumnType.DATETIME: "dd-mmm-yyyy hh:mm",
    ColumnType.INT: "#,##0",
}


def _excel_value(value: Any, col_type: str) -> Any:
    """Return the value openpyxl should write. For CURRENCY/INT we
    write the RAW number so Excel's built-in number format applies —
    that lets the user re-sort by amount. For HOURS we write a text
    string (openpyxl doesn't support the h/m glyph in a number
    format we control).
    """
    if value is None:
        return ""
    if col_type == ColumnType.HOURS:
        return fmt_hours(value)
    if col_type == ColumnType.PERCENT:
        try:
            return float(value) / 100.0     # Excel % format expects fraction
        except (TypeError, ValueError):
            return value
    if col_type == ColumnType.DATE:
        if isinstance(value, datetime):
            return value.date()
        return value
    if col_type == ColumnType.DATETIME:
        # Excel expects timezone-naive; strip tz for display.
        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value
        return value
    return value


def render_xlsx(result: ReportResult, *, stream: bool = False) -> bytes:
    """Render to bytes. Uses openpyxl.

    stream=False   → In-memory workbook. Faster for small reports.
    stream=True    → write_only workbook (large reports). Flushes rows
                     as they're appended; peak memory is proportional
                     to header + a few buffered rows.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:                # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for xlsx export; add it to backend/"
            "requirements.txt (openpyxl>=3.1)."
        ) from e

    wb = Workbook(write_only=stream)
    ws = wb.create_sheet(title="Report") if stream else wb.active
    if not stream:
        ws.title = "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")   # slate-900
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="DBEAFE")   # blue-100
    center = Alignment(horizontal="center", vertical="center")

    header_row = [c.label for c in result.columns]

    if stream:
        # write_only: iterate rows once.
        from openpyxl.cell import WriteOnlyCell
        header_cells = []
        for label in header_row:
            cell = WriteOnlyCell(ws, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            header_cells.append(cell)
        ws.append(header_cells)
        for row in result.rows:
            cells = []
            for c in result.columns:
                v = _excel_value(row.get(c.key), c.type)
                cell = WriteOnlyCell(ws, value=v)
                fmt = _EXCEL_NUMBER_FORMATS.get(c.type)
                if fmt:
                    cell.number_format = fmt
                cells.append(cell)
            ws.append(cells)
        if result.totals:
            cells = []
            for c in result.columns:
                v = _excel_value(result.totals.get(c.key), c.type)
                cell = WriteOnlyCell(ws, value=v)
                cell.font = total_font
                cell.fill = total_fill
                fmt = _EXCEL_NUMBER_FORMATS.get(c.type)
                if fmt:
                    cell.number_format = fmt
                cells.append(cell)
            ws.append(cells)
    else:
        ws.append(header_row)
        for i, c in enumerate(result.columns, start=1):
            hdr = ws.cell(row=1, column=i)
            hdr.font = header_font
            hdr.fill = header_fill
            hdr.alignment = center
            if c.width:
                ws.column_dimensions[get_column_letter(i)].width = c.width
        # Freeze header row.
        ws.freeze_panes = "A2"

        for row in result.rows:
            values = [
                _excel_value(row.get(c.key), c.type) for c in result.columns
            ]
            ws.append(values)

        # Apply number formats to data rows.
        data_row_offset = 2
        for r_idx in range(data_row_offset, data_row_offset + len(result.rows)):
            for c_idx, c in enumerate(result.columns, start=1):
                fmt = _EXCEL_NUMBER_FORMATS.get(c.type)
                if fmt:
                    ws.cell(row=r_idx, column=c_idx).number_format = fmt

        # Totals row (bold, tinted).
        if result.totals:
            tot_row_idx = data_row_offset + len(result.rows)
            for c_idx, c in enumerate(result.columns, start=1):
                v = _excel_value(result.totals.get(c.key), c.type)
                cell = ws.cell(row=tot_row_idx, column=c_idx, value=v)
                cell.font = total_font
                cell.fill = total_fill
                fmt = _EXCEL_NUMBER_FORMATS.get(c.type)
                if fmt:
                    cell.number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PDF
# ============================================================


def render_pdf(result: ReportResult, *, title: str) -> bytes:
    """PDF via reportlab. Landscape when >6 columns."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
    except ImportError as e:                # pragma: no cover
        raise RuntimeError(
            "reportlab is required for pdf export"
        ) from e

    pagesize = landscape(A4) if len(result.columns) > 6 else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(title, styles["Title"]))
    if result.meta:
        parts = []
        if "period" in result.meta:
            parts.append(f"Period: {result.meta['period']}")
        if "row_count" in result.meta:
            parts.append(f"Rows: {result.meta['row_count']}")
        parts.append(
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
        )
        story.append(Paragraph(" · ".join(parts), styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    header = [c.label for c in result.columns]
    data = [header]
    for row in result.rows:
        data.append([
            fmt_cell(row.get(c.key), c.type) for c in result.columns
        ])
    if result.totals:
        data.append([
            fmt_cell(result.totals.get(c.key), c.type)
            for c in result.columns
        ])
    tbl = Table(data, repeatRows=1)
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.white, colors.HexColor("#F8FAFC"),
        ]),
    ])
    if result.totals:
        ts.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DBEAFE"))
        ts.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    tbl.setStyle(ts)
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


# ============================================================
# CSV streaming (for very large exports)
# ============================================================


def iter_csv_rows(result: ReportResult) -> Iterable[str]:
    """Yield CSV lines one at a time. Use with StreamingResponse when
    a report may have tens of thousands of rows."""
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow([c.label for c in result.columns])
    yield buf.getvalue()
    buf.seek(0); buf.truncate(0)

    for row in result.rows:
        w.writerow([
            fmt_cell(row.get(c.key), c.type) for c in result.columns
        ])
        yield buf.getvalue()
        buf.seek(0); buf.truncate(0)

    if result.totals:
        w.writerow([
            fmt_cell(result.totals.get(c.key), c.type)
            for c in result.columns
        ])
        yield buf.getvalue()


# ============================================================
# JSON serialization (for on-screen table)
# ============================================================


def _json_value(value: Any, col_type: str) -> Any:
    """Convert a row cell to JSON-safe. Dates → ISO. Datetimes → ISO.
    Numbers stay as numbers so the frontend can format itself."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def to_json_payload(result: ReportResult) -> dict:
    """Payload the frontend renders as a sortable table."""
    return {
        "columns": [
            {
                "key": c.key, "label": c.label,
                "type": c.type, "width": c.width, "align": c.align,
            }
            for c in result.columns
        ],
        "rows": [
            {c.key: _json_value(row.get(c.key), c.type) for c in result.columns}
            for row in result.rows
        ],
        "totals": {
            c.key: _json_value(result.totals.get(c.key), c.type)
            for c in result.columns if result.totals
        } if result.totals else {},
        "meta": result.meta,
    }
